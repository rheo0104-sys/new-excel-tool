"""
DreamCIS EDC User Management Tool  v3.0
Optimizations vs v2:
  - Canvas-based grid: no widget-per-cell overhead (단일 Canvas로 전체 렌더)
  - _redraw() 에 after() debounce 적용 → 연속 키입력 중 불필요한 재렌더 방지
  - load_workbook read_only=True 로 파일 읽기, 쓸때만 일반 open
  - openpyxl import 지연 (실제 저장/로드 시점까지 미룸)
  - re 패턴 미리 컴파일 (norm, find_col 호출마다 재컴파일 방지)
  - find_col 결과 캐싱 (_col_cache) → 헤더 비교 반복 제거
  - PatternFill 오브젝트 모듈 수준 상수로 한 번만 생성
  - copy.deepcopy → dict/list 직접 복사로 교체 (inactivate 경로)
  - win32com lazy import 유지 (시작 속도 유지)
  - ttk import 제거 (사용처 없음 확인)
  - get_column_letter import 제거 (미사용)
"""

import tkinter as tk
from tkinter import messagebox, filedialog
import os, sys, re, json, datetime, glob

# ── THEME ──────────────────────────────────────────────────
PRIMARY   = "#2AACE2"
SECONDARY = "#2DC5A2"
DARK      = "#2C3E50"
LIGHT_BG  = "#F0F7FC"
HEADER_BG = "#1D8DC4"
ACCENT    = "#E8F6FC"
GREY_ROW  = "#D6D6D6"
YELLOW    = "#FFF59D"
WHITE     = "#FFFFFF"
EXISTING  = "#F0F7FB"
DISABLED  = "#E8E8E6"

SETTINGS_FILE = "dcis_settings.json"

# ── PRE-COMPILED PATTERNS ──────────────────────────────────
_RE_NORM       = re.compile(r'[\s_\-\n\r]+')
_RE_TYPE_KR    = re.compile(r'type.*계정|계정.*type')
_RE_DATE8      = re.compile(r'\d{4}[-_]?\d{2}[-_]?\d{2}')
_RE_DATE_CHECK = re.compile(r'\d{4}-\d{2}-\d{2}')

def norm(s: str) -> str:
    return _RE_NORM.sub('', str(s or "")).lower()

# ── HEADER MATCHERS ────────────────────────────────────────
def _hm(*kws):
    """Return cached header-match function for given keywords."""
    nkws = [norm(k) for k in kws]
    def match(h): return all(k in norm(h) for k in nkws)
    return match

isDateH     = _hm("date")
isTypeH     = lambda h: bool(_RE_TYPE_KR.search(norm(h)))
isEmailH    = _hm("emailaddress")
isServerH   = _hm("servertype")
isRoleH     = lambda h: norm(h).startswith("role")
isSiteNameH = _hm("sitename")
isSiteCodeH = _hm("sitecode")

# ── COLUMN LOOKUP WITH CACHE ───────────────────────────────
_col_cache: dict = {}

def find_col(headers, *kws):
    key = (id(headers), kws)
    if key in _col_cache:
        return _col_cache[key]
    nkws = [norm(k) for k in kws]
    result = next((i for i, h in enumerate(headers)
                   if all(k in norm(h) for k in nkws)), None)
    _col_cache[key] = result
    return result

def clear_col_cache():
    _col_cache.clear()

# ── EXCEL FILLS (created once) ─────────────────────────────
def _make_fills():
    from openpyxl.styles import PatternFill, Font, Alignment
    return (PatternFill("solid", fgColor="FFF59D"),   # yellow
            PatternFill("solid", fgColor="D6D6D6"),   # grey
            PatternFill("solid", fgColor="1D8DC4"),   # header blue
            Font(bold=True, color="FFFFFF"),           # header font
            Alignment(wrap_text=True))                 # header align

# ── SETTINGS ───────────────────────────────────────────────
def get_base() -> str:
    return os.path.dirname(sys.executable if getattr(sys, 'frozen', False)
                           else os.path.abspath(__file__))

def load_settings() -> dict:
    p = os.path.join(get_base(), SETTINGS_FILE)
    try:
        if os.path.exists(p):
            with open(p, encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def save_settings(d: dict):
    p = os.path.join(get_base(), SETTINGS_FILE)
    with open(p, 'w', encoding='utf-8') as f:
        json.dump(d, f, ensure_ascii=False, indent=2)

def latest_excel() -> str | None:
    base = get_base()
    files = glob.glob(os.path.join(base, "*.xlsx")) + \
            glob.glob(os.path.join(base, "*.xlsm"))
    return max(files, key=os.path.getmtime) if files else None

def make_save_path(src: str) -> str:
    today = datetime.date.today().strftime("%Y%m%d")
    d = os.path.dirname(src)
    n, ext = os.path.splitext(os.path.basename(src))
    m = _RE_DATE8.search(n)
    base_name = (n[:m.start()] + today) if m else f"{n}_{today}"
    path = os.path.join(d, base_name + ext)
    if os.path.exists(path):
        i = 1
        while os.path.exists(os.path.join(d, f"{base_name}({i}){ext}")):
            i += 1
        path = os.path.join(d, f"{base_name}({i}){ext}")
    return path

# ── BUILT-IN CALENDAR ──────────────────────────────────────
class CalendarPopup(tk.Toplevel):
    def __init__(self, parent, future_only=False, callback=None, anchor_widget=None):
        super().__init__(parent)
        self.overrideredirect(True)
        self.callback = callback
        self.future_only = future_only
        today = datetime.date.today()
        self.year, self.month = today.year, today.month
        self.selected = None
        self.configure(bg=WHITE, relief='solid', bd=1)
        self._build()
        self._position(anchor_widget or parent)
        self.grab_set()
        self.bind("<Escape>", lambda e: self.destroy())
        self.bind("<FocusOut>", lambda e: self.destroy())

    def _position(self, widget):
        self.update_idletasks()
        try:
            x = widget.winfo_rootx()
            y = widget.winfo_rooty() + widget.winfo_height() + 2
            sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
            w, h = self.winfo_width(), self.winfo_height()
            if x + w > sw: x = sw - w - 4
            if y + h > sh: y = widget.winfo_rooty() - h - 2
            self.geometry(f"+{x}+{y}")
        except Exception:
            pass

    def _build(self):
        for w in self.winfo_children(): w.destroy()
        hf = tk.Frame(self, bg=PRIMARY); hf.pack(fill='x')
        tk.Button(hf, text="◀", bg=PRIMARY, fg=WHITE, relief='flat',
                  font=("Segoe UI", 9, "bold"), bd=0, padx=6,
                  command=self._prev).pack(side='left')
        tk.Label(hf, text=f"{self.year}.{self.month:02d}", bg=PRIMARY, fg=WHITE,
                 font=("Segoe UI", 9, "bold"), width=10, anchor='center').pack(side='left', expand=True)
        tk.Button(hf, text="▶", bg=PRIMARY, fg=WHITE, relief='flat',
                  font=("Segoe UI", 9, "bold"), bd=0, padx=6,
                  command=self._next).pack(side='right')
        df = tk.Frame(self, bg=WHITE); df.pack(fill='x', padx=4, pady=(4, 0))
        for i, d in enumerate(["Su", "Mo", "Tu", "We", "Th", "Fr", "Sa"]):
            tk.Label(df, text=d, width=3, bg=WHITE,
                     fg="#999" if i in (0, 6) else DARK,
                     font=("Segoe UI", 8, "bold")).grid(row=0, column=i)
        cf = tk.Frame(self, bg=WHITE); cf.pack(fill='both', padx=4, pady=4)
        today = datetime.date.today()
        first = (datetime.date(self.year, self.month, 1).weekday() + 1) % 7
        dim = (datetime.date(self.year, self.month % 12 + 1, 1) -
               datetime.timedelta(days=1)).day if self.month < 12 else 31
        col = row = 0
        for day in range(first):
            tk.Label(cf, text="", width=3, bg=WHITE).grid(row=0, column=day)
        col = first
        for day in range(1, dim + 1):
            d = datetime.date(self.year, self.month, day)
            is_t = (d == today); is_p = self.future_only and d <= today; is_s = self.selected == day
            bg = PRIMARY if is_s else ("#D6EFFC" if is_t else WHITE)
            fg = WHITE if is_s else ("#ccc" if is_p else DARK)
            tk.Button(cf, text=str(day), width=3, bg=bg, fg=fg, relief='flat',
                      font=("Segoe UI", 8, "bold" if is_t or is_s else "normal"),
                      state='disabled' if is_p else 'normal',
                      command=lambda d=day: self._pick(d)).grid(row=row, column=col, padx=1, pady=1)
            col += 1
            if col == 7: col, row = 0, row + 1
        bf = tk.Frame(self, bg=WHITE); bf.pack(fill='x', padx=4, pady=(0, 4))
        tk.Button(bf, text="Select", bg=PRIMARY, fg=WHITE, relief='flat',
                  padx=10, pady=3, font=("Segoe UI", 8, "bold"),
                  command=self._confirm).pack(side='left', padx=2)
        tk.Button(bf, text="Cancel", bg="#E0E0E0", fg=DARK, relief='flat',
                  padx=10, pady=3, font=("Segoe UI", 8),
                  command=self.destroy).pack(side='left', padx=2)

    def _prev(self):
        self.month -= 1
        if self.month < 1: self.month, self.year = 12, self.year - 1
        self._build()

    def _next(self):
        self.month += 1
        if self.month > 12: self.month, self.year = 1, self.year + 1
        self._build()

    def _pick(self, day):
        self.selected = day; self._build()

    def _confirm(self):
        if self.selected:
            if self.callback:
                self.callback(f"{self.year}-{self.month:02d}-{self.selected:02d}")
            self.destroy()

# ── MULTI-SELECT POPUP ─────────────────────────────────────
class MultiSelectPopup(tk.Toplevel):
    def __init__(self, parent, options, current=None, callback=None):
        super().__init__(parent)
        self.title("Select Role(s)"); self.resizable(False, False)
        self.grab_set(); self.configure(bg=WHITE)
        self.callback = callback
        cur = {c.strip() for c in (current or "").split(",") if c.strip()}
        self.vars = {}
        tk.Frame(self, bg=PRIMARY, height=32).pack(fill='x')
        tk.Label(self, text="  Select Role(s)", bg=PRIMARY, fg=WHITE,
                 font=("Segoe UI", 9, "bold")).place(x=0, y=4)
        f = tk.Frame(self, bg=WHITE, padx=12, pady=8); f.pack(fill='both', expand=True)
        all_var = tk.BooleanVar()
        tk.Checkbutton(f, text="All", variable=all_var, bg=WHITE, fg=PRIMARY,
                       font=("Segoe UI", 9, "bold"), activebackground=ACCENT,
                       selectcolor=PRIMARY).pack(anchor='w', pady=(0, 4))
        for opt in options:
            v = tk.BooleanVar(value=(opt.strip() in cur))
            self.vars[opt.strip()] = v
            tk.Checkbutton(f, text=opt.strip(), variable=v, bg=WHITE, fg=DARK,
                           font=("Segoe UI", 9), activebackground=ACCENT,
                           selectcolor=PRIMARY).pack(anchor='w')
        all_var.trace_add('write', lambda *_: [v.set(all_var.get()) for v in self.vars.values()])
        bf = tk.Frame(f, bg=WHITE); bf.pack(pady=(8, 0))
        tk.Button(bf, text="OK", bg=PRIMARY, fg=WHITE, relief='flat',
                  padx=12, pady=4, font=("Segoe UI", 9, "bold"),
                  command=self._ok).pack(side='left', padx=4)
        tk.Button(bf, text="Cancel", bg="#E0E0E0", fg=DARK, relief='flat',
                  padx=12, pady=4, font=("Segoe UI", 9),
                  command=self.destroy).pack(side='left', padx=4)
        self._center(self.master)

    def _center(self, p):
        self.update_idletasks()
        px = p.winfo_rootx() + p.winfo_width() // 2
        py = p.winfo_rooty() + p.winfo_height() // 2
        w, h = self.winfo_width(), self.winfo_height()
        self.geometry(f"+{max(0,px-w//2)}+{max(0,py-h//2)}")

    def _ok(self):
        sel = [k for k, v in self.vars.items() if v.get()]
        if self.callback: self.callback(", ".join(sel))
        self.destroy()

# ── INACTIVATE DIALOG ──────────────────────────────────────
class InactivateDialog(tk.Toplevel):
    def __init__(self, parent, row_vals, headers, callback=None):
        super().__init__(parent)
        self.title("Inactivate Account"); self.resizable(False, False)
        self.grab_set(); self.configure(bg=WHITE)
        self.callback = callback
        role_idx  = find_col(headers, "role")
        name_idx  = find_col(headers, "name")
        email_idx = find_col(headers, "email", "address")
        roles_raw = (row_vals[role_idx] if role_idx is not None else "") or ""
        self.roles = [r.strip() for r in roles_raw.split(",") if r.strip()]
        name  = row_vals[name_idx]  if name_idx  is not None else ""
        email = row_vals[email_idx] if email_idx is not None else ""
        hf = tk.Frame(self, bg="#E53935", height=52)
        hf.pack(fill='x'); hf.pack_propagate(False)
        tk.Label(hf, text="  🔴  Inactivate Account",
                 bg="#E53935", fg=WHITE, font=("Segoe UI", 11, "bold")).pack(side='left', pady=12)
        content = tk.Frame(self, bg=WHITE, padx=18, pady=14); content.pack(fill='both', expand=True)
        tk.Label(content, text=f"{name}  |  {email}",
                 bg=WHITE, fg="#888", font=("Segoe UI", 8)).pack(anchor='w', pady=(0, 10))
        # Step 1
        self.step1 = tk.Frame(content, bg=WHITE); self.step1.pack(fill='both', expand=True)
        tk.Label(self.step1, text="Select role(s) to inactivate:",
                 bg=WHITE, fg=DARK, font=("Segoe UI", 9, "bold")).pack(anchor='w', pady=(0, 6))
        self.role_vars = {}
        all_var = tk.BooleanVar()
        tk.Checkbutton(self.step1, text="All Roles", variable=all_var,
                       bg="#FEE2E2", fg="#E53935", relief='solid', bd=1,
                       font=("Segoe UI", 9, "bold"), selectcolor="#E53935",
                       padx=8, pady=4, activebackground="#FEE2E2").pack(fill='x', pady=2)
        for r in self.roles:
            v = tk.BooleanVar(); self.role_vars[r] = v
            tk.Checkbutton(self.step1, text=r, variable=v, bg=WHITE, fg=DARK,
                           font=("Segoe UI", 9), selectcolor=PRIMARY).pack(anchor='w', padx=4)
        all_var.trace_add('write', lambda *_: [v.set(all_var.get()) for v in self.role_vars.values()])
        bf1 = tk.Frame(self.step1, bg=WHITE); bf1.pack(pady=(10, 0))
        tk.Button(bf1, text="Next →", bg="#E53935", fg=WHITE, relief='flat',
                  padx=12, pady=5, font=("Segoe UI", 9, "bold"),
                  command=self._to_step2).pack(side='left', padx=4)
        tk.Button(bf1, text="Cancel", bg="#E0E0E0", fg=DARK, relief='flat',
                  padx=12, pady=5, command=self.destroy).pack(side='left', padx=4)
        # Step 2
        self.step2 = tk.Frame(content, bg=WHITE)
        tk.Label(self.step2, text="Inactivation Date:", bg=WHITE, fg=DARK,
                 font=("Segoe UI", 9, "bold")).pack(anchor='w', pady=(0, 6))
        self.date_var = tk.StringVar(value="Click to select...")
        date_btn = tk.Button(self.step2, textvariable=self.date_var,
                             bg="#FFF3F3", fg="#E53935", relief='solid', bd=1,
                             font=("Segoe UI", 9), padx=10, pady=6, anchor='w')
        date_btn.config(command=lambda: CalendarPopup(self, future_only=True,
                        callback=self.date_var.set, anchor_widget=date_btn))
        date_btn.pack(fill='x', pady=4)
        bf2 = tk.Frame(self.step2, bg=WHITE); bf2.pack(pady=(10, 0))
        tk.Button(bf2, text="← Back", bg="#E0E0E0", fg=DARK, relief='flat',
                  padx=10, pady=5, command=self._to_step1).pack(side='left', padx=4)
        tk.Button(bf2, text="Confirm", bg="#E53935", fg=WHITE, relief='flat',
                  padx=12, pady=5, font=("Segoe UI", 9, "bold"),
                  command=self._confirm).pack(side='left', padx=4)
        tk.Button(bf2, text="Cancel", bg="#E0E0E0", fg=DARK, relief='flat',
                  padx=10, pady=5, command=self.destroy).pack(side='left', padx=4)
        self._center(parent)

    def _center(self, p):
        self.update_idletasks()
        px = p.winfo_rootx() + p.winfo_width() // 2
        py = p.winfo_rooty() + p.winfo_height() // 2
        w, h = self.winfo_width(), self.winfo_height()
        self.geometry(f"+{max(0,px-w//2)}+{max(0,py-h//2)}")

    def _to_step2(self):
        sel = [r for r, v in self.role_vars.items() if v.get()]
        if not sel:
            messagebox.showwarning("Selection", "Please select at least one role."); return
        self._sel_roles = sel
        self.step1.pack_forget(); self.step2.pack(fill='both', expand=True)

    def _to_step1(self):
        self.step2.pack_forget(); self.step1.pack(fill='both', expand=True)

    def _confirm(self):
        date = self.date_var.get()
        if not _RE_DATE_CHECK.match(date):
            messagebox.showwarning("Date", "Please select a date."); return
        if self.callback: self.callback(self._sel_roles, date)
        self.destroy()

# ── USER SETTINGS DIALOG ───────────────────────────────────
class UserSettingsDialog(tk.Toplevel):
    def __init__(self, parent, settings, callback=None):
        super().__init__(parent)
        self.title("User Information Settings"); self.resizable(False, False)
        self.grab_set(); self.configure(bg=WHITE)
        self.callback = callback
        hf = tk.Frame(self, bg=PRIMARY, height=52)
        hf.pack(fill='x'); hf.pack_propagate(False)
        tk.Label(hf, text="  ⚙  User Information Settings",
                 bg=PRIMARY, fg=WHITE, font=("Segoe UI", 11, "bold")).pack(side='left', pady=12)
        content = tk.Frame(self, bg=WHITE, padx=20, pady=16); content.pack(fill='both', expand=True)
        self.entries = {}
        for label, key in [("Study / Protocol Name", "study"), ("Name", "name"),
                            ("Your Email Address", "email"), ("DCIS Contact Email", "dcis_email")]:
            tk.Label(content, text=label, bg=WHITE, fg="#888",
                     font=("Segoe UI", 8, "bold")).pack(anchor='w', pady=(6, 2))
            e = tk.Entry(content, font=("Segoe UI", 9), width=42, relief='solid', bd=1)
            e.insert(0, settings.get(key, "")); e.pack(fill='x')
            self.entries[key] = e
        tk.Label(content, text="CC List  (one per line)", bg=WHITE,
                 fg="#888", font=("Segoe UI", 8, "bold")).pack(anchor='w', pady=(8, 2))
        self.cc_text = tk.Text(content, width=42, height=4, font=("Segoe UI", 9), relief='solid', bd=1)
        self.cc_text.insert('1.0', settings.get('cc', '')); self.cc_text.pack(fill='x')
        bf = tk.Frame(content, bg=WHITE); bf.pack(pady=(14, 0))
        tk.Button(bf, text="Save", bg=PRIMARY, fg=WHITE, relief='flat',
                  padx=14, pady=5, font=("Segoe UI", 9, "bold"),
                  command=self._save).pack(side='left', padx=4)
        tk.Button(bf, text="Reset", bg="#E53935", fg=WHITE, relief='flat',
                  padx=12, pady=5, command=self._reset).pack(side='left', padx=4)
        tk.Button(bf, text="Cancel", bg="#E0E0E0", fg=DARK, relief='flat',
                  padx=12, pady=5, command=self.destroy).pack(side='left', padx=4)
        self._center(parent)

    def _center(self, p):
        self.update_idletasks()
        px = p.winfo_rootx() + p.winfo_width() // 2
        py = p.winfo_rooty() + p.winfo_height() // 2
        w, h = self.winfo_width(), self.winfo_height()
        self.geometry(f"+{max(0,px-w//2)}+{max(0,py-h//2)}")

    def _save(self):
        data = {k: e.get().strip() for k, e in self.entries.items()}
        data['cc'] = self.cc_text.get('1.0', 'end-1c').strip()
        if self.callback: self.callback(data)
        self.destroy()

    def _reset(self):
        for e in self.entries.values(): e.delete(0, 'end')
        self.cc_text.delete('1.0', 'end')

# ── SHEET GRID (ttk.Treeview — C-level virtualized, handles 1000+ rows) ──────
class SheetGrid(tk.Frame):
    ROW_TAG_EXIST    = 'exist'
    ROW_TAG_NEW      = 'new'
    ROW_TAG_INPUT    = 'input'
    ROW_TAG_GREY     = 'grey'
    ROW_TAG_DISABLED = 'disabled'

    def __init__(self, master, headers, col_types, dropdown_opts,
                 readonly_cols=None, is_user=False, site_rows_ref=None, **kw):
        super().__init__(master, bg=WHITE, **kw)
        self.headers        = headers
        self.col_types      = col_types
        self.dropdown_opts  = dropdown_opts
        self.readonly_cols  = readonly_cols or set()
        self.is_user        = is_user
        self.site_rows_ref  = site_rows_ref
        self._rows: list    = []            # list of row dicts
        self._iid_to_ri: dict = {}          # treeview iid → row index
        self._filter_vals   = [""] * len(headers)
        self._edit_cell     = None          # (iid, ci)
        self._edit_widget   = None
        self._filter_frame  = None
        self._build_ui()

    # ── Build ttk.Treeview UI ──────────────────────────────
    def _build_ui(self):
        from tkinter import ttk
        style = ttk.Style(self)
        style.theme_use('default')
        style.configure("Grid.Treeview",
                        background=WHITE, foreground=DARK,
                        fieldbackground=WHITE, rowheight=22,
                        font=("Segoe UI", 9))
        style.configure("Grid.Treeview.Heading",
                        background=HEADER_BG, foreground=WHITE,
                        font=("Segoe UI", 9, "bold"), relief='flat')
        style.map("Grid.Treeview.Heading", background=[('active', PRIMARY)])
        style.map("Grid.Treeview",
                  background=[('selected', '#C5E4F5')],
                  foreground=[('selected', DARK)])

        # Column IDs
        cols = [f"c{i}" for i in range(len(self.headers))]

        container = tk.Frame(self, bg=WHITE); container.pack(fill='both', expand=True)
        vsb = ttk.Scrollbar(container, orient='vertical')
        hsb = ttk.Scrollbar(self, orient='horizontal')

        self.tv = ttk.Treeview(container, columns=cols, show='headings',
                               style="Grid.Treeview",
                               yscrollcommand=vsb.set,
                               xscrollcommand=hsb.set)
        vsb.config(command=self.tv.yview)
        hsb.config(command=self.tv.xview)
        vsb.pack(side='right', fill='y')
        self.tv.pack(side='left', fill='both', expand=True)
        hsb.pack(fill='x')

        # Configure columns & headings
        col_w = [190 if 'email' in norm(h) else
                 95  if isDateH(h)          else
                 125 if isRoleH(h)          else
                 160 if 'note' in norm(h)   else
                 100 for h in self.headers]
        for i, (cid, hdr, w) in enumerate(zip(cols, self.headers, col_w)):
            display = hdr.replace('\n', ' ')
            if self._filter_vals[i]:
                display = "▼ " + display
            self.tv.heading(cid, text=display,
                            command=lambda c=i: self._filter_click(c))
            self.tv.column(cid, width=w, minwidth=50, stretch=False)

        # Row colour tags
        self.tv.tag_configure(self.ROW_TAG_EXIST,    background=EXISTING)
        self.tv.tag_configure(self.ROW_TAG_NEW,      background=YELLOW)
        self.tv.tag_configure(self.ROW_TAG_INPUT,    background=WHITE)
        self.tv.tag_configure(self.ROW_TAG_GREY,     background=GREY_ROW, foreground="#777")
        self.tv.tag_configure(self.ROW_TAG_DISABLED, background="#E8E8E6", foreground="#BBBBBB")

        # Events
        self.tv.bind("<ButtonPress-1>",   self._on_click)
        self.tv.bind("<Double-Button-1>", self._on_dblclick)
        self.tv.bind("<Button-3>",        self._on_right_click)
        self.tv.bind("<Return>",          self._on_return)

    # ── Row tag helper ─────────────────────────────────────
    def _row_tag(self, row):
        if row.get('grey'):        return self.ROW_TAG_GREY
        if row.get('is_disabled'): return self.ROW_TAG_DISABLED
        if row.get('is_new'):      return self.ROW_TAG_NEW
        if row.get('is_input'):    return self.ROW_TAG_NEW if row.get('has_data') else self.ROW_TAG_INPUT
        return self.ROW_TAG_EXIST

    # ── Load existing rows from Excel ─────────────────────
    def load_rows(self, rows_data):
        """rows_data: list of list[str]"""
        self._rows = []
        for vals in rows_data:
            n = len(self.headers)
            v = list(vals) + [""] * max(0, n - len(vals))
            self._rows.append({'vals': v[:n], 'is_existing': True,
                               'modified_cells': set(), 'grey': False})
        # Mark greyed pairs (User sheet: same email, Active+Inactive)
        if self.is_user:
            self._recheck_grey_pairs()
        # Add the always-open input row + disabled placeholders
        self._rows.append({'vals': [""] * len(self.headers),
                           'is_input': True, 'has_data': False,
                           'modified_cells': set()})
        for _ in range(4):
            self._rows.append({'vals': [""] * len(self.headers),
                               'is_disabled': True, 'modified_cells': set()})
        self._rebuild_tree()
        # Scroll to bottom to show last existing rows
        self.after(80, lambda: self.tv.yview_moveto(1.0))

    # ── Rebuild Treeview from _rows ───────────────────────
    def _rebuild_tree(self):
        """Full rebuild — only called on load/filter change, NOT on every edit."""
        self.tv.delete(*self.tv.get_children())
        self._iid_to_ri = {}
        for ri, row in enumerate(self._rows):
            if not self._matches_filter(row):
                continue
            vals = [str(v or "") for v in row['vals']]
            tag  = self._row_tag(row)
            iid  = self.tv.insert("", "end", values=vals, tags=(tag,))
            self._iid_to_ri[iid] = ri
        self._refresh_headings()

    def _refresh_headings(self):
        cols = [f"c{i}" for i in range(len(self.headers))]
        for i, (cid, hdr) in enumerate(zip(cols, self.headers)):
            display = hdr.replace('\n', ' ')
            if self._filter_vals[i]:
                display = "▼ " + display
            self.tv.heading(cid, text=display)

    # ── Update a single row in-place (fast, no full rebuild) ─
    def _update_row_in_tree(self, iid, row):
        vals = [str(v or "") for v in row['vals']]
        tag  = self._row_tag(row)
        self.tv.item(iid, values=vals, tags=(tag,))

    # ── Filter ─────────────────────────────────────────────
    def _matches_filter(self, row):
        if row.get('is_disabled') or row.get('is_input') or row.get('is_new'):
            return True
        return all(not fv or fv.lower() in str(row['vals'][ci] or "").lower()
                   for ci, fv in enumerate(self._filter_vals))

    def _filter_click(self, ci):
        """Open a tiny filter entry popup under the heading."""
        if self._filter_frame:
            try: self._filter_frame.destroy()
            except Exception: pass
        ff = tk.Toplevel(self)
        ff.overrideredirect(True); ff.configure(bg=WHITE, relief='solid', bd=1)
        self._filter_frame = ff
        # Position under heading
        ff.update_idletasks()
        x = self.tv.winfo_rootx() + sum(self.tv.column(f"c{j}", 'width') for j in range(ci))
        y = self.tv.winfo_rooty()
        ff.geometry(f"160x52+{x}+{y}")
        tk.Label(ff, text=f"Filter: {self.headers[ci].replace(chr(10),' ')}",
                 bg=ACCENT, fg=DARK, font=("Segoe UI", 7), padx=4).pack(fill='x')
        var = tk.StringVar(value=self._filter_vals[ci])
        e = tk.Entry(ff, textvariable=var, font=("Segoe UI", 9), relief='flat', bg=WHITE)
        e.pack(fill='x', padx=3, pady=3); e.focus_set()
        def apply(*_):
            self._filter_vals[ci] = var.get().strip()
            self._rebuild_tree()
            try: ff.destroy()
            except Exception: pass
        e.bind("<Return>",  apply)
        e.bind("<Escape>",  lambda _: ff.destroy())
        ff.bind("<FocusOut>", lambda _: ff.destroy())

    # ── Events ─────────────────────────────────────────────
    def _iid_ri(self, iid):
        return self._iid_to_ri.get(iid)

    def _on_click(self, event):
        iid = self.tv.identify_row(event.y)
        if not iid: return
        ri  = self._iid_ri(iid)
        if ri is None: return
        row = self._rows[ri]
        if row.get('is_disabled'):
            self.tv.yview_moveto(1.0); return
        col_id = self.tv.identify_column(event.x)   # '#1', '#2', ...
        ci = int(col_id[1:]) - 1 if col_id else 0
        if row.get('is_input') or row.get('is_new'):
            self._open_edit(iid, ri, ci)

    def _on_dblclick(self, event):
        iid = self.tv.identify_row(event.y)
        if not iid: return
        ri  = self._iid_ri(iid)
        if ri is None: return
        row = self._rows[ri]
        if row.get('is_disabled') or row.get('grey'): return
        if not row.get('is_input') and not row.get('is_new'):
            col_id = self.tv.identify_column(event.x)
            ci = int(col_id[1:]) - 1 if col_id else 0
            self._open_edit(iid, ri, ci)

    def _on_right_click(self, event):
        if not self.is_user: return
        iid = self.tv.identify_row(event.y)
        if not iid: return
        ri  = self._iid_ri(iid)
        if ri is None: return
        row = self._rows[ri]
        if not row.get('is_existing') or row.get('grey'): return
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="🔴  Inactivate Account",
                         command=lambda: self._do_inactivate(ri))
        menu.tk_popup(event.x_root, event.y_root)

    def _on_return(self, event):
        sel = self.tv.selection()
        if sel: self._on_dblclick(type('E', (), {'y': self.tv.bbox(sel[0])[1]+1, 'x': 5})())

    # ── Edit cell ──────────────────────────────────────────
    def _open_edit(self, iid, ri, ci):
        if self._edit_widget:
            try: self._edit_widget.destroy()
            except Exception: pass
        row = self._rows[ri]
        ct  = self.col_types[ci] if ci < len(self.col_types) else 'text'
        val = str(row['vals'][ci] or "")
        # Email readonly for existing rows
        if isEmailH(self.headers[ci]) and row.get('is_existing'): return
        # Type col always readonly
        if ct == 'type': return

        if ct in ('date', 'date_future'):
            CalendarPopup(self, future_only=(ct == 'date_future'),
                          callback=lambda v: self._commit(iid, ri, ci, v),
                          anchor_widget=self)
            return
        if ct == 'dropdown':
            opts = self.dropdown_opts.get(ci, [])
            if callable(opts): opts = opts(ri, ci, row)
            DropdownPopup(self, opts, current=val,
                          callback=lambda v: self._commit(iid, ri, ci, v),
                          anchor_widget=self)
            return
        if ct == 'multi':
            opts = self.dropdown_opts.get(ci, [])
            if callable(opts): opts = opts(ri, ci, row)
            MultiSelectPopup(self, opts, current=val,
                             callback=lambda v: self._commit(iid, ri, ci, v))
            return

        # Inline text entry overlay
        bbox = self.tv.bbox(iid, column=f"c{ci}")
        if not bbox: return
        x, y, w, h = bbox
        var = tk.StringVar(value=val)
        e = tk.Entry(self.tv, textvariable=var, font=("Segoe UI", 9),
                     relief='solid', bd=1, bg=YELLOW)
        if isEmailH(self.headers[ci]):
            var.trace_add('write', lambda *_: var.set(var.get().replace(' ', '')))
        e.place(x=x, y=y, width=w, height=h)
        e.focus_set(); e.select_range(0, 'end')
        self._edit_widget = e
        self._edit_cell   = (iid, ri, ci)

        def commit(*_):
            self._commit(iid, ri, ci, var.get())
            try: e.destroy()
            except Exception: pass
        def tab_next(shift=False):
            commit()
            n = len(self.headers)
            nci = (ci - 1) if shift else (ci + 1)
            if 0 <= nci < n:
                self.after(10, lambda: self._open_edit(iid, ri, nci))
        e.bind("<Return>",            lambda _: commit())
        e.bind("<Escape>",            lambda _: e.destroy())
        e.bind("<Tab>",               lambda _: tab_next(False))
        e.bind("<Shift-Tab>",         lambda _: tab_next(True))
        e.bind("<FocusOut>",          lambda _: commit())

    def _commit(self, iid, ri, ci, new_val):
        """Write value, update row state, refresh just this row in Treeview."""
        row     = self._rows[ri]
        old_val = str(row['vals'][ci] or "")
        nv      = str(new_val or "")

        row['vals'][ci] = nv

        # Server type change → clear site name/code
        if self.is_user and isServerH(self.headers[ci]):
            sni = next((i for i, h in enumerate(self.headers) if isSiteNameH(h)), None)
            sci = next((i for i, h in enumerate(self.headers) if isSiteCodeH(h)), None)
            if sni is not None: row['vals'][sni] = ""
            if sci is not None: row['vals'][sci] = ""

        # Site name ↔ code sync
        if self.is_user and self.site_rows_ref:
            srv = row['vals'][next((i for i,h in enumerate(self.headers) if isServerH(h)), -1)] if any(isServerH(h) for h in self.headers) else ""
            site_rows = self.site_rows_ref()
            if isSiteNameH(self.headers[ci]):
                m = next((r for r in site_rows if r['vals'][1] == nv and (not srv or r['vals'][0].lower() == srv.lower())), None)
                if m:
                    ski = next((i for i,h in enumerate(self.headers) if isSiteCodeH(h)), None)
                    if ski is not None: row['vals'][ski] = m['vals'][2]
            elif isSiteCodeH(self.headers[ci]):
                m = next((r for r in site_rows if r['vals'][2] == nv and (not srv or r['vals'][0].lower() == srv.lower())), None)
                if m:
                    sni = next((i for i,h in enumerate(self.headers) if isSiteNameH(h)), None)
                    if sni is not None: row['vals'][sni] = m['vals'][1]
            # All Team / ALL
            if isSiteNameH(self.headers[ci]) and nv == "All Team":
                ski = next((i for i,h in enumerate(self.headers) if isSiteCodeH(h)), None)
                if ski is not None: row['vals'][ski] = "ALL"
            if isSiteCodeH(self.headers[ci]) and nv == "ALL":
                sni = next((i for i,h in enumerate(self.headers) if isSiteNameH(h)), None)
                if sni is not None: row['vals'][sni] = "All Team"

        # Track modified cells
        if row.get('is_existing') and nv != old_val:
            row.setdefault('modified_cells', set()).add(ci)
        elif row.get('is_existing') and nv == old_val:
            row.get('modified_cells', set()).discard(ci)

        # Input row: first data → mark has_data, activate next disabled row, auto Active
        if row.get('is_input') and not row.get('has_data') and any(v for v in row['vals']):
            row['has_data'] = True
            # Auto Active for type col
            if self.is_user:
                ti = next((i for i,h in enumerate(self.headers) if isTypeH(h)), None)
                if ti is not None and not row['vals'][ti]:
                    row['vals'][ti] = "Active"
            # Activate the first disabled row as new input row
            for r2 in self._rows:
                if r2.get('is_disabled'):
                    r2.clear(); r2.update({'vals': [""] * len(self.headers),
                                           'is_input': True, 'has_data': False,
                                           'modified_cells': set()})
                    # Insert new input row into treeview
                    new_iid = self.tv.insert("", "end",
                                             values=[""] * len(self.headers),
                                             tags=(self.ROW_TAG_INPUT,))
                    self._iid_to_ri[new_iid] = self._rows.index(r2)
                    break

        # Refresh just this one row — no full redraw!
        self._update_row_in_tree(iid, row)
        self._edit_widget = None
        self._edit_cell   = None

    # ── Temp save (site sheet) ─────────────────────────────
    def temp_save(self) -> int:
        """Promote all input rows with data to is_new. Return count promoted."""
        n = 0
        for iid, ri in list(self._iid_to_ri.items()):
            row = self._rows[ri]
            if row.get('is_input') and row.get('has_data'):
                row['is_input'] = False; row['is_new'] = True
                row.setdefault('modified_cells', set())
                self._update_row_in_tree(iid, row)
                n += 1
        return n

    # ── Badge count ───────────────────────────────────────
    def badge_count(self) -> int:
        mod  = sum(1 for r in self._rows if r.get('is_existing') and r.get('modified_cells'))
        new  = sum(1 for r in self._rows if r.get('is_new') and not r.get('grey'))
        inact= sum(1 for r in self._rows if r.get('is_new') and r.get('grey') and
                   'inactive' in str(r['vals'][1] if len(r['vals']) > 1 else "").lower())
        return mod + new + inact

    def get_all_rows(self): return self._rows

    def site_lookup_rows(self):
        return [r for r in self._rows
                if (r.get('is_existing') or r.get('is_new')) and not r.get('is_disabled')]

    # ── Grey pair detection ────────────────────────────────
    def _recheck_grey_pairs(self):
        ei = next((i for i,h in enumerate(self.headers) if isEmailH(h)), None)
        ti = next((i for i,h in enumerate(self.headers) if isTypeH(h)), None)
        ai = next((i for i,h in enumerate(self.headers) if isDateH(h)), None)
        ni = next((i for i,h in enumerate(self.headers) if 'note' in norm(h)), None)
        if ei is None: return
        from collections import defaultdict
        by_email = defaultdict(list)
        for ri, r in enumerate(self._rows):
            if r.get('is_existing') or r.get('is_new'):
                em = str(r['vals'][ei] or "").lower().strip()
                if em: by_email[em].append(ri)
        for em, idxs in by_email.items():
            types = {str(self._rows[i]['vals'][ti] or "").lower() for i in idxs} if ti is not None else set()
            if 'active' in types and 'inactive' in types:
                act_date   = next((str(self._rows[i]['vals'][ai] or "")
                                   for i in idxs if str(self._rows[i]['vals'][ti] or "").lower() == 'active'), "")
                inact_date = next((str(self._rows[i]['vals'][ai] or "")
                                   for i in idxs if 'inactive' in str(self._rows[i]['vals'][ti] or "").lower()), "")
                note_txt = f"계정 만료\nActivation: {act_date} / Inactivation: {inact_date}"
                for i in idxs:
                    self._rows[i]['grey'] = True
                    if ni is not None:
                        existing_note = str(self._rows[i]['vals'][ni] or "")
                        if note_txt not in existing_note:
                            self._rows[i]['vals'][ni] = (existing_note + "\n" + note_txt).strip()

    # ── Inactivate ────────────────────────────────────────
    def _do_inactivate(self, ri):
        row = self._rows[ri]
        InactivateDialog(self, row['vals'], self.headers,
                         callback=lambda d: self._apply_inactivate(ri, d))

    def _apply_inactivate(self, ri, data):
        inact_roles = data['roles']; inact_date = data['date']
        row    = self._rows[ri]
        ri_idx = find_col(self.headers, "role")
        ti_idx = find_col(self.headers, "type")
        ai_idx = find_col(self.headers, "activation", "date")
        ni_idx = next((i for i,h in enumerate(self.headers) if 'note' in norm(h)), None)
        if ri_idx is None: return
        all_roles  = [r.strip() for r in str(row['vals'][ri_idx] or "").split(",") if r.strip()]
        rem_roles  = [r for r in all_roles if r not in inact_roles]
        inact_str  = ", ".join(inact_roles)
        act_date   = str(row['vals'][ai_idx] or "") if ai_idx is not None else ""
        note_txt   = f"계정 만료\nActivation: {act_date} / Inactivation: {inact_date}"
        is_all     = len(rem_roles) == 0

        import copy

        def make_inact_row():
            v = row['vals'][:]
            if ti_idx is not None: v[ti_idx] = "Inactive"
            if ai_idx is not None: v[ai_idx] = inact_date
            if ri_idx is not None: v[ri_idx] = inact_str
            if ni_idx is not None: v[ni_idx] = note_txt
            return {'vals': v, 'is_existing': True, 'is_new': True,
                    'grey': True, 'modified_cells': set()}

        if is_all:
            # In-place: grey + note on original
            row['grey'] = True
            if ni_idx is not None:
                existing = str(row['vals'][ni_idx] or "")
                row['vals'][ni_idx] = (existing + "\n" + note_txt).strip() if existing else note_txt
        else:
            # Partial: update original role only (no colour change)
            row['vals'][ri_idx] = ", ".join(rem_roles)
            # Split-active row right after original
            split_vals = row['vals'][:]
            split_vals[ri_idx] = inact_str
            if ni_idx is not None: split_vals[ni_idx] = note_txt
            split_row = {'vals': split_vals, 'is_existing': True, 'is_new': True,
                         'grey': True, 'modified_cells': set()}
            # Insert after ri (before input/disabled)
            insert_at = ri + 1
            self._rows.insert(insert_at, split_row)

        # Inactive row always at end (before input/disabled)
        first_special = next((i for i, r in enumerate(self._rows)
                               if r.get('is_input') or r.get('is_disabled')), len(self._rows))
        self._rows.insert(first_special, make_inact_row())
        self._rebuild_tree()

    # ── Dropdown popup (simple listbox popup) ─────────────
    # (reused by _open_edit above via DropdownPopup class)

    def _build_ui(self):
        self.hdr_canvas = tk.Canvas(self, height=42, bg=HEADER_BG, highlightthickness=0)
        self.hdr_canvas.pack(fill='x')
        self.hdr_canvas.bind("<ButtonPress-1>",   self._on_hdr_press)
        self.hdr_canvas.bind("<B1-Motion>",        self._on_hdr_drag)
        self.hdr_canvas.bind("<ButtonRelease-1>",  self._on_hdr_release)
        body = tk.Frame(self, bg=WHITE); body.pack(fill='both', expand=True)
        from tkinter import ttk
        vsb = ttk.Scrollbar(body, orient='vertical')
        hsb = ttk.Scrollbar(self, orient='horizontal')
        self.body_canvas = tk.Canvas(body, bg=WHITE, highlightthickness=0,
                                     yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        def _vsb_cmd(*args):
            self.body_canvas.yview(*args)
            self._schedule_redraw()   # re-virtualize after scrollbar drag
        vsb.config(command=_vsb_cmd)
        hsb.config(command=self.body_canvas.xview)
        vsb.pack(side='right', fill='y')
        self.body_canvas.pack(side='left', fill='both', expand=True)
        hsb.pack(fill='x')
        self.body_canvas.bind("<Configure>",       self._on_canvas_resize)
        self.body_canvas.bind("<MouseWheel>",      self._on_mousewheel)
        self.body_canvas.bind("<ButtonPress-1>",   self._on_body_click)
        self.body_canvas.bind("<Double-Button-1>", self._on_body_dblclick)
        self.body_canvas.bind("<Button-3>",        self._on_right_click)
        self.hdr_canvas.bind("<Configure>",        self._draw_header)

    def _total_width(self):
        return sum(self._col_widths) + 2

    # ── Header ─────────────────────────────────────────────
    def _draw_header(self, _event=None):
        c = self.hdr_canvas; c.delete("all"); x = 0
        for i, (h, w) in enumerate(zip(self.headers, self._col_widths)):
            c.create_rectangle(x, 0, x+w, 42, fill=HEADER_BG, outline="")
            c.create_rectangle(x, 38, x+w, 42, fill=PRIMARY, outline="")
            if self._filter_vals[i]:
                c.create_rectangle(x+w-18, 4, x+w-4, 18, fill="#FFF59D", outline=PRIMARY)
                c.create_text(x+w-11, 11, text="▼", fill=DARK, font=("Segoe UI", 7))
            else:
                c.create_rectangle(x+w-18, 4, x+w-4, 18, fill="", outline="white")
                c.create_text(x+w-11, 11, text="▼", fill="white", font=("Segoe UI", 7))
            c.create_text(x+6, 20, text=h.replace('\n', ' '), anchor='w',
                          fill=WHITE, font=("Segoe UI", 8, "bold"), width=w-24)
            c.create_line(x+w-1, 0, x+w-1, 42, fill="white")
            x += w
        c.config(scrollregion=(0, 0, self._total_width(), 42))

    # ── Body (debounced + virtualized) ─────────────────────
    def _schedule_redraw(self):
        if not self._redraw_pending:
            self._redraw_pending = True
            self.after(self.REDRAW_DELAY, self._redraw)

    def _redraw(self):
        self._redraw_pending = False
        if self._edit_widget:
            try: self._edit_widget.destroy()
            except Exception: pass
            self._edit_widget = None
        c = self.body_canvas

        # ── Pass 1: compute ALL row Y positions (no drawing) ──────
        # This is cheap (just arithmetic) and must cover all rows
        # so hit-testing (_row_at_y) works correctly.
        y = 0; self._row_y = []
        total_h = 0
        for row in self._rows:
            if not self._matches_filter(row):
                self._row_y.append(None); continue
            self._row_y.append(y)
            rh = self.ROW_H_INPUT if (row.get('is_input') or row.get('is_new')) else self.ROW_H_EXIST
            y += rh
        total_h = y

        # Update scrollregion BEFORE reading viewport
        c.config(scrollregion=(0, 0, self._total_width(), max(total_h, 1)))

        # ── Pass 2: draw only rows visible in viewport (true virtualization) ─
        c.delete("all")
        canvas_h = c.winfo_height() or 600
        # canvasy(0) = top of visible area in canvas coords
        view_top    = int(c.canvasy(0))
        view_bottom = view_top + canvas_h + self.ROW_H_EXIST  # small overdraw buffer

        for ri, row in enumerate(self._rows):
            ry = self._row_y[ri]
            if ry is None: continue
            rh = self.ROW_H_INPUT if (row.get('is_input') or row.get('is_new')) else self.ROW_H_EXIST
            # Skip rows completely outside viewport
            if ry + rh < view_top: continue
            if ry > view_bottom:   break   # rows are in order → safe to stop
            x = 0
            for ci, w in enumerate(self._col_widths):
                bg = self._cell_bg(row, ci)
                c.create_rectangle(x, ry, x+w, ry+rh, fill=bg, outline="#E4EFF8")
                val = str(row['vals'][ci] or "")
                ct = self.col_types[ci] if ci < len(self.col_types) else 'text'
                if ct in ('dropdown', 'multi') and not row.get('grey') and not row.get('is_disabled'):
                    c.create_text(x+w-10, ry+rh//2, text="▾",
                                  fill=PRIMARY if ct == 'dropdown' else SECONDARY,
                                  font=("Segoe UI", 9, "bold"))
                    max_w = w - 22
                else:
                    max_w = w - 8
                tf = DARK
                if isTypeH(self.headers[ci]):
                    tf = SECONDARY if val.lower() == 'active' else "#E53935"
                elif row.get('grey'):
                    tf = "#777"
                c.create_text(x+4, ry+4, text=val, anchor='nw', fill=tf,
                              font=("Segoe UI", 8, "bold" if isTypeH(self.headers[ci]) else "normal"),
                              width=max_w)
                x += w

        self._draw_header()
        # Scroll to bottom on initial load only (flag cleared after first use)
        if getattr(self, '_scroll_to_bottom', False):
            self._scroll_to_bottom = False
            self.after(50, lambda: self.body_canvas.yview_moveto(1.0))

    def _cell_bg(self, row, ci):
        if row.get('grey'):        return GREY_ROW
        if row.get('is_disabled'): return DISABLED
        if row.get('is_new'):      return YELLOW
        if row.get('is_input'):    return YELLOW if row.get('has_data') else WHITE
        if row.get('is_existing'):
            return YELLOW if ci in row.get('modified_cells', set()) else EXISTING
        return WHITE

    def _repaint_row(self, ri):
        """Fast single-row repaint: delete only this row's canvas items and redraw."""
        ry = self._row_y[ri] if ri < len(self._row_y) else None
        if ry is None:
            # Row not currently visible or position not yet computed → full redraw
            self._schedule_redraw(); return
        row = self._rows[ri]
        rh  = self.ROW_H_INPUT if (row.get('is_input') or row.get('is_new')) else self.ROW_H_EXIST
        c   = self.body_canvas
        # Delete only items overlapping this row's y-band
        for item in c.find_overlapping(0, ry, self._total_width(), ry + rh):
            c.delete(item)
        # Redraw just this row
        x = 0
        for ci, w in enumerate(self._col_widths):
            bg = self._cell_bg(row, ci)
            c.create_rectangle(x, ry, x+w, ry+rh, fill=bg, outline="#E4EFF8")
            val = str(row['vals'][ci] or "")
            ct  = self.col_types[ci] if ci < len(self.col_types) else 'text'
            if ct in ('dropdown', 'multi') and not row.get('grey') and not row.get('is_disabled'):
                c.create_text(x+w-10, ry+rh//2, text="▾",
                              fill=PRIMARY if ct == 'dropdown' else SECONDARY,
                              font=("Segoe UI", 9, "bold"))
                max_w = w - 22
            else:
                max_w = w - 8
            tf = DARK
            if isTypeH(self.headers[ci]):
                tf = SECONDARY if val.lower() == 'active' else "#E53935"
            elif row.get('grey'): tf = "#777"
            c.create_text(x+4, ry+4, text=val, anchor='nw', fill=tf,
                          font=("Segoe UI", 8, "bold" if isTypeH(self.headers[ci]) else "normal"),
                          width=max_w)
            x += w

    def _matches_filter(self, row):
        if row.get('is_disabled') or row.get('is_input') or row.get('is_new'): return True
        return all(not fv or fv.lower() in str(row['vals'][ci] or "").lower()
                   for ci, fv in enumerate(self._filter_vals))

    # ── Hit-testing ────────────────────────────────────────
    def _row_at_y(self, y):
        for ri, ry in enumerate(self._row_y):
            if ry is None: continue
            rh = (self.ROW_H_INPUT if (self._rows[ri].get('is_input') or
                  self._rows[ri].get('is_new')) else self.ROW_H_EXIST)
            if ry <= y < ry + rh: return ri
        return None

    def _col_at_x(self, x):
        cx = 0
        for ci, w in enumerate(self._col_widths):
            if cx <= x < cx + w: return ci
            cx += w
        return None

    def _canvas_coords(self, event):
        return self.body_canvas.canvasx(event.x), self.body_canvas.canvasy(event.y)

    # ── Events ─────────────────────────────────────────────
    def _on_body_click(self, event):
        x, y = self._canvas_coords(event)
        ri = self._row_at_y(y)
        if ri is None: return
        row = self._rows[ri]; ci = self._col_at_x(x)
        if ci is None: return
        if row.get('is_disabled'):
            self.body_canvas.yview_moveto(1.0); return
        if row.get('is_input') or row.get('is_new'):
            self._open_edit(ri, ci)

    def _on_body_dblclick(self, event):
        x, y = self._canvas_coords(event)
        ri = self._row_at_y(y)
        if ri is None: return
        row = self._rows[ri]; ci = self._col_at_x(x)
        if ci is None: return
        if row.get('is_disabled') or row.get('grey'): return
        if not row.get('is_input') and not row.get('is_new'):
            self._open_edit(ri, ci)

    def _on_right_click(self, event):
        if not self.is_user: return
        x, y = self._canvas_coords(event)
        ri = self._row_at_y(y)
        if ri is None: return
        row = self._rows[ri]
        if not row.get('is_existing') or row.get('grey'): return
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="🔴  Inactivate Account",
                         command=lambda: self._do_inactivate(ri))
        menu.tk_popup(event.x_root, event.y_root)

    def _on_mousewheel(self, event):
        self.body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        # Re-draw visible rows after scroll (virtualization)
        self._schedule_redraw()

    def _on_canvas_resize(self, _event):
        self._draw_header()
        self._schedule_redraw()  # viewport size changed → re-virtualize

    # ── Column resize ──────────────────────────────────────
    def _on_hdr_press(self, event):
        x = self.hdr_canvas.canvasx(event.x); cx = 0
        for ci, w in enumerate(self._col_widths):
            if abs(x - (cx + w)) < 6:
                self._drag_col = ci; self._drag_x = event.x; self._drag_w = w; return
            cx += w
        cx = 0
        for ci, w in enumerate(self._col_widths):
            if cx + w - 18 <= x <= cx + w - 4 and event.y <= 18:
                self._open_filter(ci, event); return
            cx += w

    def _on_hdr_drag(self, event):
        if self._drag_col is None: return
        self._col_widths[self._drag_col] = max(self.MIN_COL_W,
                                               self._drag_w + event.x - self._drag_x)
        self._schedule_redraw()

    def _on_hdr_release(self, _event):
        self._drag_col = None

    # ── Filter ─────────────────────────────────────────────
    def _open_filter(self, ci, event):
        pop = tk.Toplevel(self); pop.overrideredirect(True)
        pop.configure(bg=WHITE, relief='solid', bd=1)
        x = self.hdr_canvas.winfo_rootx() + sum(self._col_widths[:ci])
        y = self.hdr_canvas.winfo_rooty() + 42
        pop.geometry(f"180x70+{x}+{y}")
        var = tk.StringVar(value=self._filter_vals[ci])
        tk.Label(pop, text=f"Filter: {self.headers[ci].split(chr(10))[0]}",
                 bg=WHITE, fg=DARK, font=("Segoe UI", 8)).pack(anchor='w', padx=6, pady=2)
        e = tk.Entry(pop, textvariable=var, font=("Segoe UI", 9), relief='solid', bd=1)
        e.pack(fill='x', padx=6); e.focus()
        def apply(*_):
            self._filter_vals[ci] = var.get(); self._schedule_redraw(); pop.destroy()
        e.bind('<Return>', apply); e.bind('<Escape>', lambda ev: pop.destroy())

    # ── Cell editing ───────────────────────────────────────
    def _open_edit(self, ri, ci):
        row = self._rows[ri]
        if isTypeH(self.headers[ci]): return
        if isEmailH(self.headers[ci]) and row.get('is_existing'): return
        if ci in self.readonly_cols: return
        ct = self.col_types[ci] if ci < len(self.col_types) else 'text'
        if ct in ('date', 'date_future'):
            ry = self._row_y[ri]
            if ry is None: return
            bx = self.body_canvas.winfo_rootx() + sum(self._col_widths[:ci]) - int(self.body_canvas.canvasx(0))
            by = self.body_canvas.winfo_rooty() + ry - int(self.body_canvas.canvasy(0))
            class _A:
                def winfo_rootx(_): return bx
                def winfo_rooty(_): return by
                def winfo_height(_): return SheetGrid.ROW_H_EXIST
            CalendarPopup(self.winfo_toplevel(), future_only=(ct == 'date_future'),
                          callback=lambda d: self._commit(ri, ci, d), anchor_widget=_A())
            return
        if ct == 'multi':
            opts = self.dropdown_opts.get(ci, [])
            if callable(opts): opts = opts(ri, ci)
            MultiSelectPopup(self.winfo_toplevel(), opts,
                             current=str(row['vals'][ci] or ""),
                             callback=lambda v: self._commit(ri, ci, v))
            return
        if ct == 'dropdown':
            opts = self.dropdown_opts.get(ci, [])
            if callable(opts): opts = opts(ri, ci)
            self._show_dropdown(ri, ci, opts); return
        self._show_entry(ri, ci)

    def _show_entry(self, ri, ci):
        if self._edit_widget:
            try: self._edit_widget.destroy()
            except Exception: pass
        row = self._rows[ri]; ry = self._row_y[ri]
        if ry is None: return
        rh = self.ROW_H_INPUT if (row.get('is_input') or row.get('is_new')) else self.ROW_H_EXIST
        rx  = sum(self._col_widths[:ci]) - int(self.body_canvas.canvasx(0))
        rys = ry - int(self.body_canvas.canvasy(0))
        e = tk.Entry(self.body_canvas, font=("Segoe UI", 8), relief='solid', bd=2,
                     highlightthickness=1, highlightcolor=PRIMARY, highlightbackground=PRIMARY)
        e.insert(0, str(row['vals'][ci] or "")); e.place(x=rx, y=rys, width=self._col_widths[ci], height=rh)
        e.focus(); e.select_range(0, 'end')
        self._edit_widget = e; self._edit_cell = (ri, ci)
        def commit(_ev=None):
            val = e.get()
            if isEmailH(self.headers[ci]): val = val.replace(" ", "")
            self._commit(ri, ci, val)
        def tab_next(ev):
            commit(); self._open_edit(ri, (ci + 1) % len(self.headers)); return "break"
        def tab_prev(ev):
            commit(); self._open_edit(ri, (ci - 1) % len(self.headers)); return "break"
        e.bind('<Return>', commit); e.bind('<Tab>', tab_next)
        e.bind('<Shift-Tab>', tab_prev)
        e.bind('<Escape>', lambda ev: (e.destroy(), setattr(self, '_edit_widget', None)))
        e.bind('<FocusOut>', commit)

    def _show_dropdown(self, ri, ci, opts):
        if self._edit_widget:
            try: self._edit_widget.destroy()
            except Exception: pass
        row = self._rows[ri]; ry = self._row_y[ri]
        if ry is None: return
        rh = self.ROW_H_INPUT if (row.get('is_input') or row.get('is_new')) else self.ROW_H_EXIST
        rx  = sum(self._col_widths[:ci]) - int(self.body_canvas.canvasx(0))
        rys = ry - int(self.body_canvas.canvasy(0))
        pop = tk.Toplevel(self); pop.overrideredirect(True)
        pop.configure(bg=WHITE, relief='solid', bd=1)
        pop.geometry(f"+{self.body_canvas.winfo_rootx()+rx}+{self.body_canvas.winfo_rooty()+rys+rh}")
        self._edit_widget = pop; self._edit_cell = (ri, ci)
        for opt in opts:
            bg = ACCENT if opt == str(row['vals'][ci] or "") else WHITE
            b = tk.Button(pop, text=opt, bg=bg, fg=DARK, relief='flat',
                          font=("Segoe UI", 9), anchor='w', padx=10, pady=4,
                          command=lambda o=opt: (self._commit(ri, ci, o), pop.destroy()))
            b.pack(fill='x')
            b.bind('<Enter>', lambda ev, btn=b: btn.config(bg=ACCENT))
            b.bind('<Leave>', lambda ev, btn=b, o=opt:
                   btn.config(bg=ACCENT if o == str(row['vals'][ci] or "") else WHITE))
        pop.bind('<FocusOut>', lambda ev: pop.destroy()); pop.focus()

    # ── Commit ─────────────────────────────────────────────
    def _commit(self, ri, ci, val):
        if self._edit_widget:
            try: self._edit_widget.destroy()
            except Exception: pass
            self._edit_widget = None
        row = self._rows[ri]
        old_val = str(row['vals'][ci] or "")
        val = str(val or "")
        row['vals'][ci] = val
        # Server type change → clear site name/code
        if self.is_user and isServerH(self.headers[ci]):
            for fn in (lambda h: find_col(self.headers, "site", "name"),
                       lambda h: find_col(self.headers, "site", "code")):
                idx = fn(None)
                if idx is not None: row['vals'][idx] = ""
        # Site name ↔ code sync
        if self.is_user:
            srv = (row['vals'][2] or "").lower()
            site_rows = self.site_rows_ref() if callable(self.site_rows_ref) else []
            if isSiteNameH(self.headers[ci]):
                m = next((r for r in site_rows if r['vals'][1] == val and
                          (not srv or r['vals'][0].lower() == srv)), None)
                if m:
                    ki = find_col(self.headers, "site", "code")
                    if ki is not None: row['vals'][ki] = m['vals'][2]
            if isSiteCodeH(self.headers[ci]):
                m = next((r for r in site_rows if r['vals'][2] == val and
                          (not srv or r['vals'][0].lower() == srv)), None)
                if m:
                    ni = find_col(self.headers, "site", "name")
                    if ni is not None: row['vals'][ni] = m['vals'][1]
            if isSiteNameH(self.headers[ci]) and val == "All Team":
                ki = find_col(self.headers, "site", "code")
                if ki is not None: row['vals'][ki] = "ALL"
            if isSiteCodeH(self.headers[ci]) and val == "ALL":
                ni = find_col(self.headers, "site", "name")
                if ni is not None: row['vals'][ni] = "All Team"
        # Track modified cells (existing rows, only on actual change)
        if row.get('is_existing'):
            mc = row.setdefault('modified_cells', set())
            if val != old_val: mc.add(ci)
            else:              mc.discard(ci)
        # Input row: hasData + auto Active + activate next disabled row
        if row.get('is_input'):
            had = row.get('has_data', False)
            row['has_data'] = any(v for v in row['vals'])
            if self.is_user and row['has_data']:
                ti = find_col(self.headers, "type", "계정") or find_col(self.headers, "type")
                if ti is not None and not row['vals'][ti]:
                    row['vals'][ti] = "Active"
            if row['has_data'] and not had:
                my = self._rows.index(row)
                for j in range(my + 1, len(self._rows)):
                    if self._rows[j].get('is_disabled'):
                        self._rows[j]['is_disabled'] = False
                        self._rows[j]['is_input'] = True
                        break
                # Row state changed (new input row opened) → need full redraw
                self._schedule_redraw(); return

        # For simple value changes: just repaint the single visible row
        # instead of redrawing the entire canvas
        self._repaint_row(ri)

    # ── Inactivate ─────────────────────────────────────────
    def _do_inactivate(self, ri):
        row = self._rows[ri]
        def on_confirm(sel_roles, date):
            type_idx = find_col(self.headers, "type", "계정") or find_col(self.headers, "type")
            date_idx = find_col(self.headers, "date")
            role_idx = find_col(self.headers, "role")
            note_idx = find_col(self.headers, "note")
            all_r    = [r.strip() for r in (row['vals'][role_idx] or "").split(",") if r.strip()]
            remain   = [r for r in all_r if r not in sel_roles]
            act_date = row['vals'][date_idx] if date_idx is not None else ""
            note_txt = f"계정 만료\nActivation: {act_date} / Inactivation: {date}"
            is_all   = (len(remain) == 0)
            # helper: shallow-copy row dict with new vals list
            def clone(src): return {**src, 'vals': src['vals'][:], 'modified_cells': set()}
            if is_all:
                row['grey'] = True; row['modified_cells'] = set()
                if note_idx is not None:
                    row['vals'][note_idx] = (row['vals'][note_idx] or "").rstrip()
                    if row['vals'][note_idx]: row['vals'][note_idx] += "\n"
                    row['vals'][note_idx] += note_txt
            else:
                if role_idx is not None: row['vals'][role_idx] = ", ".join(remain)
                row['modified_cells'] = set()
                split = clone(row)
                split['grey'] = True
                if role_idx is not None: split['vals'][role_idx] = ", ".join(sel_roles)
                if note_idx is not None: split['vals'][note_idx] = note_txt
                self._rows.insert(self._rows.index(row) + 1, split)
            inact = clone(row)
            inact['grey'] = True
            if type_idx is not None: inact['vals'][type_idx] = "Inactive"
            if date_idx is not None: inact['vals'][date_idx] = date
            if role_idx is not None: inact['vals'][role_idx] = ", ".join(sel_roles)
            if note_idx is not None: inact['vals'][note_idx] = note_txt
            first = next((j for j, r in enumerate(self._rows)
                          if r.get('is_input') or r.get('is_disabled')), len(self._rows))
            self._rows.insert(first, inact)
            self._scroll_to_bottom = False  # don't jump to bottom after inactivate
            self._schedule_redraw()
        InactivateDialog(self.winfo_toplevel(), row['vals'], self.headers, callback=on_confirm)

    # ── Data load ──────────────────────────────────────────
    def load_rows(self, data_rows, n_disabled=4):
        self._rows = []
        for vals in data_rows:
            v = [str(x) if x is not None else "" for x in vals]
            v += [""] * max(0, len(self.headers) - len(v))
            self._rows.append({'vals': v[:len(self.headers)], 'is_existing': True,
                               'is_new': False, 'is_input': False, 'is_disabled': False,
                               'modified_cells': set(), 'grey': False, 'has_data': False})
        if self.is_user: self._mark_grey_pairs()
        self._rows.append(self._new_input_row())
        for _ in range(n_disabled):
            self._rows.append({'vals': [""] * len(self.headers), 'is_existing': False,
                               'is_new': False, 'is_input': False, 'is_disabled': True,
                               'modified_cells': set(), 'grey': False, 'has_data': False})
        self._scroll_to_bottom = True   # scroll once after first render
        self._schedule_redraw()

    def _mark_grey_pairs(self):
        ei = find_col(self.headers, "email", "address")
        ti = find_col(self.headers, "type", "계정") or find_col(self.headers, "type")
        if ei is None or ti is None: return
        groups: dict = {}
        for r in self._rows:
            em = (r['vals'][ei] or "").strip().lower()
            if em: groups.setdefault(em, []).append(r)
        for rows in groups.values():
            types = {(r['vals'][ti] or "").lower() for r in rows}
            if 'active' in types and 'inactive' in types:
                for r in rows: r['grey'] = True

    def _new_input_row(self):
        return {'vals': [""] * len(self.headers), 'is_existing': False,
                'is_new': False, 'is_input': True, 'is_disabled': False,
                'modified_cells': set(), 'grey': False, 'has_data': False}

    def temp_save(self):
        count = 0
        for row in self._rows:
            if row.get('is_input') and row.get('has_data'):
                row['is_input'] = False; row['is_new'] = True; count += 1
        if not any(r.get('is_input') for r in self._rows):
            fd = next((i for i, r in enumerate(self._rows) if r.get('is_disabled')), len(self._rows))
            self._rows.insert(fd, self._new_input_row())
        self._schedule_redraw(); return count

    def get_all_rows(self): return self._rows

    def badge_count(self):
        mod        = len({id(r) for r in self._rows if r.get('is_existing') and r.get('modified_cells')})
        new_active = sum(1 for r in self._rows if r.get('is_new') and not r.get('grey'))
        new_inact  = sum(1 for r in self._rows if r.get('is_new') and r.get('grey'))
        return mod + new_active + new_inact

    def site_lookup_rows(self):
        return [r for r in self._rows if r.get('is_existing') or r.get('is_new')]

# ── MAIN APPLICATION ───────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("DreamCIS EDC User Management")
        self.geometry("1280x820"); self.minsize(960, 640)
        self.configure(bg=LIGHT_BG)
        self.settings = load_settings()
        self.cur_file = None
        self.wb       = None
        self._site_hdr = []; self._user_hdr = []
        self._site_r2  = []; self._user_r2  = []
        self._build_ui()
        self._auto_load()

    def _build_ui(self):
        self._build_topbar(); self._build_filebar()
        self._build_tabs(); self._build_bottombar()

    def _build_topbar(self):
        bar = tk.Frame(self, bg=PRIMARY, height=56)
        bar.pack(fill='x'); bar.pack_propagate(False)
        lf = tk.Frame(bar, bg=PRIMARY); lf.pack(side='left', padx=14)
        tk.Label(lf, text="dream", bg=PRIMARY, fg=WHITE, font=("Segoe UI", 18, "bold")).pack(side='left')
        tk.Label(lf, text="cis", bg=PRIMARY, fg=SECONDARY, font=("Segoe UI", 18, "bold")).pack(side='left')
        tk.Label(lf, text=" A TIGERMED COMPANY", bg=PRIMARY, fg="#aaddf5", font=("Segoe UI", 7)).pack(side='left', pady=(8, 0))
        tk.Frame(bar, bg="white", width=1).pack(side='left', fill='y', padx=12, pady=8)
        inf = tk.Frame(bar, bg=PRIMARY); inf.pack(side='left')
        self._lbl_study = tk.Label(inf, text="", bg=PRIMARY, fg=WHITE, font=("Segoe UI", 9, "bold")); self._lbl_study.pack(anchor='w')
        sf = tk.Frame(inf, bg=PRIMARY); sf.pack(anchor='w')
        self._lbl_name = tk.Label(sf, text="Name: —", bg=PRIMARY, fg="#aaddf5", font=("Segoe UI", 8)); self._lbl_name.pack(side='left')
        tk.Label(sf, text="  |  ", bg=PRIMARY, fg="#aaddf5", font=("Segoe UI", 8)).pack(side='left')
        self._lbl_dcis = tk.Label(sf, text="DCIS: —", bg=PRIMARY, fg="#aaddf5", font=("Segoe UI", 8)); self._lbl_dcis.pack(side='left')
        tk.Button(bar, text="⚙  Settings", bg=SECONDARY, fg=WHITE, relief='flat',
                  padx=12, pady=5, font=("Segoe UI", 9, "bold"),
                  command=self._open_settings,
                  activebackground="#26b090").pack(side='right', padx=14, pady=10)
        self._refresh_labels()

    def _refresh_labels(self):
        s = self.settings
        self._lbl_study.config(text=s.get('study', ''))
        self._lbl_name.config(text=f"Name: {s.get('name', '—')}")
        self._lbl_dcis.config(text=f"DCIS: {s.get('dcis_email', '—')}")

    def _open_settings(self):
        UserSettingsDialog(self, self.settings,
                           callback=lambda d: (self.settings.update(d),
                                               save_settings(self.settings),
                                               self._refresh_labels()))

    def _build_filebar(self):
        bar = tk.Frame(self, bg=WHITE, highlightbackground="#D5EAF5", highlightthickness=1)
        bar.pack(fill='x')
        inner = tk.Frame(bar, bg=WHITE); inner.pack(fill='x', padx=12, pady=6)
        self._lbl_file = tk.Label(inner, text="No file loaded", bg=WHITE, fg=DARK, font=("Segoe UI", 9))
        self._lbl_file.pack(side='left', padx=6)
        tk.Button(inner, text="Change File", bg=ACCENT, fg=PRIMARY, relief='flat',
                  padx=10, pady=3, font=("Segoe UI", 8, "bold"), bd=1,
                  command=self._browse_file).pack(side='right')

    def _build_tabs(self):
        outer = tk.Frame(self, bg=LIGHT_BG); outer.pack(fill='both', expand=True, padx=8, pady=(8, 0))
        tab_bar = tk.Frame(outer, bg=LIGHT_BG); tab_bar.pack(fill='x')
        self._active_tab = tk.StringVar(value="site"); self._tab_btns = {}
        for key, label in [("site", "Site Information_cube"), ("user", "User Information_cube")]:
            btn = tk.Button(tab_bar, text=label, relief='flat', padx=16, pady=7,
                            font=("Segoe UI", 9, "bold"), command=lambda k=key: self._switch_tab(k))
            btn.pack(side='left', padx=(0, 2)); self._tab_btns[key] = btn
        self._badge_vars = {'site': tk.StringVar(value=""), 'user': tk.StringVar(value="")}
        self._tempsave_btn = tk.Button(tab_bar, text="💾  Temp. Save",
                                       bg=SECONDARY, fg=WHITE, relief='flat', padx=12, pady=5,
                                       font=("Segoe UI", 9, "bold"), command=self._temp_save,
                                       activebackground="#26b090")
        self._tempsave_btn.pack(side='right', padx=4)
        self._panel = tk.Frame(outer, bg=WHITE, bd=1, relief='solid'); self._panel.pack(fill='both', expand=True)
        self._grids = {}; self._switch_tab("site")

    def _switch_tab(self, key):
        self._active_tab.set(key)
        for k, btn in self._tab_btns.items():
            badge = self._badge_vars[k].get()
            base = "Site Information_cube" if k == "site" else "User Information_cube"
            btn.config(text=f"{base}  [{badge}]" if badge else base,
                       bg=PRIMARY if k == key else "#CBE8F5",
                       fg=WHITE if k == key else "#4A7A95",
                       activebackground=PRIMARY)
        self._tempsave_btn.pack(side='right', padx=4) if key == "site" else self._tempsave_btn.pack_forget()
        for w in self._panel.winfo_children(): w.pack_forget()
        if key in self._grids: self._grids[key].pack(fill='both', expand=True)

    def _temp_save(self):
        if 'site' in self._grids:
            n = self._grids['site'].temp_save(); self._update_badges()
            if n: messagebox.showinfo("Temp. Save",
                                      f"{n} row(s) saved temporarily.\n"
                                      "Site codes are now available in User Information dropdowns.")

    def _update_badges(self):
        for k in ('site', 'user'):
            if k in self._grids:
                n = self._grids[k].badge_count()
                self._badge_vars[k].set(str(n) if n else "")
        self._switch_tab(self._active_tab.get())

    def _build_bottombar(self):
        bar = tk.Frame(self, bg=WHITE, highlightbackground="#D5EAF5", highlightthickness=1)
        bar.pack(fill='x')
        inner = tk.Frame(bar, bg=WHITE); inner.pack(fill='x', padx=12, pady=8)
        for color, label in [(EXISTING, "Existing (dbl-click)"), (YELLOW, "Modified / New"),
                              (GREY_ROW, "Expired pair"), (DISABLED, "Disabled")]:
            tk.Label(inner, text="■", bg=WHITE, fg=color, font=("Segoe UI", 10)).pack(side='left')
            tk.Label(inner, text=label + "  ", bg=WHITE, fg="#888", font=("Segoe UI", 8)).pack(side='left')
        for text, bg, cmd in [("Save As...", "#607D8B", self._save_as),
                               ("💾  Save", PRIMARY, self._save),
                               ("✉  Save & Send Mail", SECONDARY, self._save_and_mail)]:
            tk.Button(inner, text=text, bg=bg, fg=WHITE, relief='flat',
                      padx=14, pady=5, font=("Segoe UI", 9, "bold"),
                      command=cmd).pack(side='right', padx=4)

    # ── File loading ───────────────────────────────────────
    def _auto_load(self):
        p = self.settings.get('last_file')
        if p and os.path.exists(p): self._load(p); return
        p = latest_excel()
        if p: self._load(p)

    def _browse_file(self):
        p = filedialog.askopenfilename(initialdir=get_base(),
                                       filetypes=[("Excel", "*.xlsx *.xlsm"), ("All", "*.*")])
        if p: self._load(p)

    def _load(self, path):
        try:
            from openpyxl import load_workbook as lw
            import threading

            self.cur_file = path
            self.settings['last_file'] = path
            save_settings(self.settings)
            self._lbl_file.config(text=os.path.basename(path))
            clear_col_cache()
            if not self.settings.get('study'):
                self.settings['study'] = os.path.basename(path).split("_")[0]
                self._refresh_labels()

            # ── Fix 1: read_only=True only, never load writable wb at startup ──
            # wb is opened writable ONLY at save time (_collect_and_save)
            self.wb        = None   # will be opened on demand at save
            self._wb_path  = path   # remember path for save-time open

            def _bg_load():
                """Background thread: read data only (read_only=True = fast)."""
                try:
                    wbr   = lw(path, read_only=True, data_only=True)
                    sheets = wbr.sheetnames
                    site_sh = next((s for s in sheets if 'site' in s.lower()), sheets[0])
                    user_sh = next((s for s in sheets if 'user' in s.lower()),
                                   sheets[1] if len(sheets) > 1 else None)
                    site_data = list(wbr[site_sh].iter_rows(values_only=True)) if site_sh else []
                    user_data = list(wbr[user_sh].iter_rows(values_only=True)) if user_sh else []
                    wbr.close()
                    # Marshal results back to main thread via after()
                    self.after(0, lambda: self._on_load_done(site_sh, user_sh,
                                                              site_data, user_data))
                except Exception as e:
                    self.after(0, lambda: messagebox.showerror("Load Error", str(e)))

            self._lbl_file.config(text=f"Loading {os.path.basename(path)}…")
            threading.Thread(target=_bg_load, daemon=True).start()

        except Exception as e:
            messagebox.showerror("Load Error", str(e))

    def _on_load_done(self, site_sh, user_sh, site_data, user_data):
        """Called on the main thread after background load completes."""
        self._lbl_file.config(text=os.path.basename(self.cur_file))
        if site_sh and site_data: self._parse_site_data(site_data)
        if user_sh and user_data: self._parse_user_data(user_data)
        self._update_badges()

    def _parse_site_data(self, rows):
        if not rows: return
        self._site_hdr = [str(c or "") for c in rows[0]]
        self._site_r2  = [str(c or "") for c in rows[1]] if len(rows) > 1 else []
        data = rows[2:]
        hidden = {find_col(self._site_hdr, "request", "date"),
                  find_col(self._site_hdr, "request", "person")} - {None}
        self._site_disp_hdr = [h for i, h in enumerate(self._site_hdr) if i not in hidden]
        self._site_disp_idx = [i for i in range(len(self._site_hdr)) if i not in hidden]
        col_types, drop_opts = [], {}
        for di, oi in enumerate(self._site_disp_idx):
            h = self._site_hdr[oi]
            if isServerH(h):
                opts = [x.strip() for x in re.split(r'[/,]', self._site_r2[oi] if oi < len(self._site_r2) else "") if x.strip()]
                col_types.append('dropdown'); drop_opts[di] = opts
            elif isDateH(h): col_types.append('date')
            else:            col_types.append('text')
        disp = [[str(r[i] or "") for i in self._site_disp_idx] for r in data]
        for w in self._panel.winfo_children(): w.destroy()
        self._grids = {}
        g = SheetGrid(self._panel, self._site_disp_hdr, col_types, drop_opts, is_user=False)
        g.load_rows(disp); self._grids['site'] = g

    def _parse_user_data(self, rows):
        if not rows: return
        self._user_hdr = [str(c or "") for c in rows[0]]
        self._user_r2  = [str(c or "") for c in rows[1]] if len(rows) > 1 else []
        data = rows[2:]
        hidden = {find_col(self._user_hdr, "request", "date"),
                  find_col(self._user_hdr, "request", "person")} - {None}
        self._user_disp_hdr = [h for i, h in enumerate(self._user_hdr) if i not in hidden]
        self._user_disp_idx = [i for i in range(len(self._user_hdr)) if i not in hidden]
        col_types, drop_opts = [], {}
        for di, oi in enumerate(self._user_disp_idx):
            h = self._user_hdr[oi]
            if isTypeH(h):    col_types.append('type')
            elif isServerH(h):
                opts = [x.strip() for x in re.split(r'[/,]', self._user_r2[oi] if oi < len(self._user_r2) else "") if x.strip()]
                col_types.append('dropdown'); drop_opts[di] = opts
            elif isSiteNameH(h):
                col_types.append('dropdown')
                drop_opts[di] = lambda ri, ci, _: self._site_name_opts(ri)
            elif isSiteCodeH(h):
                col_types.append('dropdown')
                drop_opts[di] = lambda ri, ci, _: self._site_code_opts(ri)
            elif isRoleH(h):
                r2 = self._user_r2[oi] if oi < len(self._user_r2) else ""
                opts = [x.strip() for x in r2.split(",") if x.strip()]
                col_types.append('multi'); drop_opts[di] = opts or []
            elif isDateH(h):  col_types.append('date_future')
            else:             col_types.append('text')
        disp = [[str(r[i] or "") for i in self._user_disp_idx] for r in data]
        g = SheetGrid(self._panel, self._user_disp_hdr, col_types, drop_opts,
                      is_user=True,
                      site_rows_ref=lambda: (self._grids['site'].site_lookup_rows()
                                             if 'site' in self._grids else []))
        g.load_rows(disp); self._grids['user'] = g

    def _site_name_opts(self, ri):
        if 'user' not in self._grids or 'site' not in self._grids: return []
        rows = self._grids['user'].get_all_rows()
        si   = next((i for i, h in enumerate(self._grids['user'].headers) if isServerH(h)), None)
        srv  = (rows[ri]['vals'][si] if si is not None and ri < len(rows) else "").lower()
        seen, result = set(), []
        for r in self._grids['site'].site_lookup_rows():
            if not srv or r['vals'][0].lower() == srv:
                v = r['vals'][1]
                if v and v not in seen: seen.add(v); result.append(v)
        result.sort()
        if "All Team" not in result: result.append("All Team")
        return result

    def _site_code_opts(self, ri):
        if 'user' not in self._grids or 'site' not in self._grids: return []
        rows = self._grids['user'].get_all_rows()
        si   = next((i for i, h in enumerate(self._grids['user'].headers) if isServerH(h)), None)
        srv  = (rows[ri]['vals'][si] if si is not None and ri < len(rows) else "").lower()
        seen, result = set(), []
        for r in self._grids['site'].site_lookup_rows():
            if not srv or r['vals'][0].lower() == srv:
                v = r['vals'][2]
                if v and v not in seen: seen.add(v); result.append(v)
        result.sort(key=lambda x: (len(x), x))
        if "ALL" not in result: result.append("ALL")
        return result

    # ── Save ───────────────────────────────────────────────
    def _collect_and_save(self, dest):
        if not self.cur_file:
            messagebox.showwarning("No file", "No file loaded."); return False
        try:
            from openpyxl import load_workbook as lw
            from openpyxl.styles import PatternFill, Font, Alignment
            today = datetime.date.today().strftime("%Y-%m-%d")
            name  = self.settings.get('name', '')
            # Open writable copy only now (at save time) — never held open in memory
            wb    = lw(self._wb_path)
            sheets = wb.sheetnames
            site_sh = next((s for s in sheets if 'site' in s.lower()), sheets[0])
            user_sh = next((s for s in sheets if 'user' in s.lower()),
                           sheets[1] if len(sheets) > 1 else None)
            yf    = PatternFill("solid", fgColor="FFF59D")
            gf    = PatternFill("solid", fgColor="D6D6D6")
            hf    = PatternFill("solid", fgColor="1D8DC4")
            hfont = Font(bold=True, color="FFFFFF")
            halign = Alignment(wrap_text=True)

            def write_sheet(ws_name, disp_hdr, disp_idx, all_hdr, req_d, req_p, grid_rows):
                ws = wb[ws_name]
                for row in ws.iter_rows(min_row=3):
                    for cell in row: cell.value = None
                out = 3
                for rd in grid_rows:
                    if rd.get('is_disabled') or rd.get('is_input'): continue
                    full = [""] * len(all_hdr)
                    for di, oi in enumerate(disp_idx):
                        full[oi] = rd['vals'][di] if di < len(rd['vals']) else ""
                    if req_d is not None:  full[req_d] = today
                    if req_p is not None:  full[req_p] = name
                    ti = find_col(all_hdr, "type", "계정") or find_col(all_hdr, "type")
                    if rd.get('is_new') and ti is not None and not full[ti]: full[ti] = "Active"
                    for ci2, val in enumerate(full):
                        cell = ws.cell(row=out, column=ci2 + 1, value=val)
                        if rd.get('grey'): cell.fill = gf
                        elif rd.get('is_new') or (rd.get('modified_cells') and
                             any(disp_idx[k] == ci2 for k in rd.get('modified_cells', set())
                                 if k < len(disp_idx))):
                            cell.fill = yf
                    out += 1
                for cell in ws[1]:
                    cell.fill = hf; cell.font = hfont; cell.alignment = halign

            if site_sh and 'site' in self._grids:
                write_sheet(site_sh, self._site_disp_hdr, self._site_disp_idx,
                            self._site_hdr,
                            find_col(self._site_hdr, "request", "date"),
                            find_col(self._site_hdr, "request", "person"),
                            self._grids['site'].get_all_rows())
            if user_sh and 'user' in self._grids:
                write_sheet(user_sh, self._user_disp_hdr, self._user_disp_idx,
                            self._user_hdr,
                            find_col(self._user_hdr, "request", "date"),
                            find_col(self._user_hdr, "request", "person"),
                            self._grids['user'].get_all_rows())
            wb.save(dest); return True
        except Exception as e:
            messagebox.showerror("Save Error", str(e)); return False

    def _save(self):
        if not self.cur_file:
            messagebox.showwarning("No file", "No file loaded."); return
        dest = make_save_path(self.cur_file)
        if self._collect_and_save(dest):
            self.cur_file = dest; self.settings['last_file'] = dest
            save_settings(self.settings); self._lbl_file.config(text=os.path.basename(dest))
            messagebox.showinfo("Saved", f"Saved:\n{os.path.basename(dest)}")

    def _save_as(self):
        default = os.path.basename(make_save_path(self.cur_file)) if self.cur_file else "output.xlsx"
        p = filedialog.asksaveasfilename(initialdir=get_base(), initialfile=default,
                                         defaultextension=".xlsx",
                                         filetypes=[("Excel", "*.xlsx")])
        if p and self._collect_and_save(p):
            self.cur_file = p; self.settings['last_file'] = p
            save_settings(self.settings); self._lbl_file.config(text=os.path.basename(p))
            messagebox.showinfo("Saved", f"Saved:\n{os.path.basename(p)}")

    def _save_and_mail(self):
        if not self.cur_file:
            messagebox.showwarning("No file", "No file loaded."); return
        dest = make_save_path(self.cur_file)
        if not self._collect_and_save(dest): return
        self.cur_file = dest; self.settings['last_file'] = dest
        save_settings(self.settings); self._lbl_file.config(text=os.path.basename(dest))
        self._send_mail(dest)

    # ── Mail ───────────────────────────────────────────────
    def _build_mail_body(self, dest):
        today = datetime.date.today().strftime("%Y-%m-%d")
        study = self.settings.get('study', '')
        lines = [f"[{study}] User Information has been updated as follows.",
                 "", f"Date: {today}", ""]
        for key, label in [('site', 'Site Information'), ('user', 'User Information')]:
            if key not in self._grids: continue
            rows = self._grids[key].get_all_rows()
            mod       = sum(1 for r in rows if r.get('is_existing') and r.get('modified_cells'))
            new_act   = [r for r in rows if r.get('is_new') and not r.get('grey')]
            new_inact = [r for r in rows if r.get('is_new') and r.get('grey')]
            if not (mod or new_act or new_inact): continue
            hdr = self._site_disp_hdr if key == 'site' else self._user_disp_hdr
            def gc(*kws): return next((i for i, h in enumerate(hdr) if all(k in norm(h) for k in kws)), None)
            lines.append(f"[ {label} ]")
            if mod: lines.append(f"  • {mod} row(s) modified")
            for r in new_act:
                srv = r['vals'][gc("server","type")] if gc("server","type") is not None else ""
                sc  = r['vals'][gc("site","code")]   if gc("site","code")   is not None else ""
                sn  = r['vals'][gc("site","name")]   if gc("site","name")   is not None else ""
                nm  = r['vals'][gc("name")]           if gc("name")          is not None else ""
                em  = r['vals'][gc("email","address")] if gc("email","address") is not None else ""
                rl  = r['vals'][gc("role")]           if gc("role")          is not None else ""
                lines.append(f"  + [Active]   {srv} / {sc} / {sn} | {nm} <{em}> | Role: {rl}")
            for r in new_inact:
                srv = r['vals'][gc("server","type")] if gc("server","type") is not None else ""
                sc  = r['vals'][gc("site","code")]   if gc("site","code")   is not None else ""
                nm  = r['vals'][gc("name")]           if gc("name")          is not None else ""
                em  = r['vals'][gc("email","address")] if gc("email","address") is not None else ""
                rl  = r['vals'][gc("role")]           if gc("role")          is not None else ""
                ad  = r['vals'][gc("activation","date")] if gc("activation","date") is not None else ""
                lines.append(f"  - [Inactive]  {srv} / {sc} | {nm} <{em}> | Role: {rl} | Inactivation: {ad}")
            lines.append("")
        lines += ["Best regards,", self.settings.get('name', '')]
        return "\n".join(lines)

    def _send_mail(self, dest):
        study   = self.settings.get('study', '')
        to_addr = self.settings.get('dcis_email', '')
        cc_list = [e.strip() for e in self.settings.get('cc', '').splitlines() if e.strip()]
        subject = f"[{study}] User Information 이 업데이트 되었습니다."
        body    = self._build_mail_body(dest)
        try:
            import win32com.client as win32
            ol   = win32.Dispatch('Outlook.Application')
            mail = ol.CreateItem(0)
            mail.To = to_addr; mail.Subject = subject
            mail.HTMLBody = ("<pre style='font-family:Segoe UI,sans-serif;font-size:9pt'>"
                             + body.replace('\n', '<br>') + "</pre>")
            if cc_list: mail.CC = "; ".join(cc_list)
            mail.Attachments.Add(os.path.abspath(dest))
            mail.Display(); return
        except ImportError:
            pass
        except Exception as ex:
            messagebox.showwarning("Outlook", f"Outlook error: {ex}\nFalling back to mailto.")
        import urllib.parse, webbrowser
        url = (f"mailto:{urllib.parse.quote(to_addr)}"
               f"?subject={urllib.parse.quote(subject)}"
               f"&body={urllib.parse.quote(body)}")
        if cc_list: url += f"&cc={urllib.parse.quote(', '.join(cc_list))}"
        webbrowser.open(url)
        messagebox.showinfo("Mail", f"Mail draft opened.\nPlease manually attach:\n{os.path.basename(dest)}")

# ── ENTRY POINT ────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
